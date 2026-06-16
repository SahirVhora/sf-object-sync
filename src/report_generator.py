"""
Excel report generator - produces a 4-sheet workbook in output/.

Sheet 1 Summary        : run metadata and status counts
Sheet 2 Object Detail  : one row per processed object
Sheet 3 Payload Log    : one row per POST attempted
Sheet 4 Validation Errors: rows rejected in Phase 1

openpyxl formatting:
  Headers: bold, fill #1F4E79, white text
  UPLOAD_SUCCESS: light green
  UPLOAD_FAILED:  light red
  DEV_EXISTS:     light grey
  Auto-fit column widths
"""

import json
import logging
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from .audit_logger import (
    ALL_STATUSES,
    UPLOAD_SUCCESS,
    UPLOAD_FAILED,
    DEV_EXISTS,
)

logger = logging.getLogger(__name__)

# ── Colour palettes ─────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")

ROW_STATUS_FILLS = {
    UPLOAD_SUCCESS: GREEN_FILL,
    UPLOAD_FAILED: RED_FILL,
    DEV_EXISTS: GREY_FILL,
}


def _apply_header(ws, row: int, headers: List[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_fit(ws) -> None:
    """Approximate auto-fit column widths based on cell content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Cap at 60 chars to avoid extremely wide columns
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _row_fill(status: str) -> Optional[PatternFill]:
    return ROW_STATUS_FILLS.get(status)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_summary(ws, run_meta: Dict[str, Any], audit_records: List[Dict]) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    _apply_header(ws, 1, ["Field", "Value"])

    rows = [
        ("Run ID", run_meta.get("run_id", "")),
        ("Started At", run_meta.get("started_at", "")),
        ("Completed At", run_meta.get("completed_at", "")),
        ("Config File", run_meta.get("config_file", "")),
        ("Input File", run_meta.get("input_file", "")),
        ("Dry Run", str(run_meta.get("dry_run", True))),
        ("PRD Base URL", run_meta.get("prd_url", "")),
        ("Dev Base URL", run_meta.get("dev_url", "")),
        ("Total Input Rows", run_meta.get("total_rows", "")),
        ("Valid Input Rows", run_meta.get("valid_rows", "")),
        ("", ""),
        ("── Status Counts ──", ""),
    ]

    # Count each unique entity/status combination once.
    # Audit records may include the same entity in multiple phases
    # (for example gap_check and upload both log DEV_EXISTS), so raw
    # record counts would inflate summary totals.
    status_counts: Dict[str, int] = {s: 0 for s in ALL_STATUSES}
    seen: set = set()
    for rec in audit_records:
        s = rec.get("status", "")
        key = (
            rec.get("object_type", ""),
            rec.get("external_code", ""),
            s,
        )
        if s in status_counts and key not in seen:
            seen.add(key)
            status_counts[s] += 1

    for status in ALL_STATUSES:
        label = " ".join(word.capitalize() for word in status.split("_"))
        rows.append((label, status_counts[status]))

    for r_idx, (label, value) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=label)
        ws.cell(row=r_idx, column=2, value=value)


def _build_object_detail(ws, rows: List[Dict]) -> None:
    ws.title = "Object Detail"
    headers = [
        "Input Object",
        "Input Code",
        "Level",
        "Entity Set",
        "External Code",
        "PRD Status",
        "Dev Status (Before)",
        "Action Taken",
        "Upload Status",
        "Error",
        "Timestamp",
    ]
    _apply_header(ws, 1, headers)

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.get("input_object", ""),
            row.get("input_code", ""),
            row.get("level", ""),
            row.get("entity_set", ""),
            row.get("external_code", ""),
            row.get("prd_status", ""),
            row.get("dev_status_before", ""),
            row.get("action_taken", ""),
            row.get("upload_status", ""),
            row.get("error", ""),
            row.get("timestamp", ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            fill = _row_fill(row.get("upload_status", ""))
            if fill:
                cell.fill = fill

    _auto_fit(ws)


def _build_payload_log(ws, audit_records: List[Dict]) -> None:
    ws.title = "Payload Log"
    headers = [
        "External Code",
        "Entity Set",
        "Payload JSON",
        "HTTP Status",
        "Response",
        "Timestamp",
    ]
    _apply_header(ws, 1, headers)

    post_records = [
        r for r in audit_records
        if r.get("payload_sent") is not None
    ]

    for r_idx, rec in enumerate(post_records, start=2):
        payload_str = json.dumps(rec.get("payload_sent", {}))
        response_str = json.dumps(rec.get("response_received", {}))
        values = [
            rec.get("external_code", ""),
            rec.get("entity_set", ""),
            payload_str[:32767],   # Excel cell limit
            rec.get("http_status", ""),
            response_str[:1000],
            rec.get("timestamp", ""),
        ]
        status = rec.get("status", "")
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            fill = _row_fill(status)
            if fill:
                cell.fill = fill

    _auto_fit(ws)


def _build_validation_errors(ws, validation_errors: List[Dict]) -> None:
    ws.title = "Validation Errors"
    headers = ["Row #", "Input Object", "Input Code", "Reason", "Timestamp"]
    _apply_header(ws, 1, headers)

    for r_idx, err in enumerate(validation_errors, start=2):
        values = [
            err.get("row_number", ""),
            err.get("input_object", ""),
            err.get("input_code", ""),
            err.get("reason", ""),
            err.get("timestamp", ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill = RED_FILL

    _auto_fit(ws)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_report(
    output_dir: str,
    run_meta: Dict[str, Any],
    audit_records: List[Dict],
    object_detail_rows: List[Dict],
    validation_errors: List[Dict],
) -> str:
    """
    Generate the 4-sheet Excel report.

    Returns the absolute path to the written .xlsx file.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    ws1 = wb.create_sheet("Summary")
    ws2 = wb.create_sheet("Object Detail")
    ws3 = wb.create_sheet("Payload Log")
    ws4 = wb.create_sheet("Validation Errors")

    _build_summary(ws1, run_meta, audit_records)
    _build_object_detail(ws2, object_detail_rows)
    _build_payload_log(ws3, audit_records)
    _build_validation_errors(ws4, validation_errors)

    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
    filename = f"sync_report_{ts}.xlsx"
    path = os.path.join(output_dir, filename)
    wb.save(path)
    logger.info("Excel report written: %s", path)
    return path
