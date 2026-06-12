"""
tests/test_sync_engine.py - unit tests for src/sync_engine.py

Tests mock SFClient at the boundary - no live API calls are made.
"""

import os
import sys
import tempfile
import unittest
from unittest.mock import MagicMock, patch, call
from typing import Any, Dict

import openpyxl

# ── Path bootstrap ────────────────────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

from src.sync_engine import (
    sync_objects,
    config_from_env,
    _phase1_validate,
)
from src.audit_logger import AuditLogger, VALIDATION_FAILED


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_xlsx(rows: list, tmp_dir: str) -> str:
    """Write a minimal input xlsx to a temp file and return its path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Object", "Code"])
    for row in rows:
        ws.append(list(row))
    path = os.path.join(tmp_dir, "input.xlsx")
    wb.save(path)
    return path


# ── Phase 1 validation tests ──────────────────────────────────────────────────

class TestPhase1Validate(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def _audit(self):
        return AuditLogger(self.tmp, dry_run=True).__enter__()

    def test_valid_rows_returned(self):
        path = _make_xlsx([("Sub Department", "10000073"), ("Department", "DEP-001")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], {"object_type": "Sub Department", "code": "10000073"})
        self.assertEqual(len(errors), 0)

    def test_invalid_object_type_rejected(self):
        path = _make_xlsx([("Division", "DIV-001")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(len(rows), 0)
        self.assertEqual(len(errors), 1)
        self.assertIn("Invalid Object type", errors[0]["reason"])

    def test_empty_code_rejected(self):
        path = _make_xlsx([("Department", "")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(len(rows), 0)
        self.assertEqual(errors[0]["reason"], "Code is empty")

    def test_invalid_code_chars_rejected(self):
        path = _make_xlsx([("Department", "DEP@001")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(len(rows), 0)
        self.assertIn("invalid characters", errors[0]["reason"])

    def test_case_insensitive_object_type(self):
        path = _make_xlsx([("sub department", "10000001")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(rows[0]["object_type"], "Sub Department")

    def test_blank_rows_skipped(self):
        path = _make_xlsx([("", ""), ("Department", "DEP-001"), ("", "")], self.tmp)
        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)
        self.assertEqual(len(rows), 1)

    def test_missing_file_raises(self):
        audit = self._audit()
        with self.assertRaises(FileNotFoundError):
            _phase1_validate("/nonexistent/file.xlsx", audit, [])

    def test_missing_header_raises(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Foo", "Bar"])
        path = os.path.join(self.tmp, "no_header.xlsx")
        wb.save(path)
        audit = self._audit()
        with self.assertRaises(ValueError):
            _phase1_validate(path, audit, [])

    def test_header_can_be_on_non_active_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Foo", "Bar"])
        input_ws = wb.create_sheet("Input")
        input_ws.append(["Object", "Code"])
        input_ws.append(["Department", "DEP-001"])
        wb.active = 0
        path = os.path.join(self.tmp, "second_sheet_input.xlsx")
        wb.save(path)

        audit = self._audit()
        errors = []
        rows = _phase1_validate(path, audit, errors)

        self.assertEqual(rows, [{"object_type": "Department", "code": "DEP-001"}])
        self.assertEqual(errors, [])


# ── config_from_env tests ─────────────────────────────────────────────────────

class TestConfigFromEnv(unittest.TestCase):

    def test_raises_without_urls(self):
        with patch.dict(os.environ, {}, clear=True):
            # Remove SF_SOURCE_URL and SF_TARGET_URL if present
            env = {k: v for k, v in os.environ.items()
                   if k not in ("SF_SOURCE_URL", "SF_TARGET_URL")}
            with patch.dict(os.environ, env, clear=True):
                with self.assertRaises(ValueError):
                    config_from_env()

    def test_reads_env_vars(self):
        env = {
            "SF_SOURCE_URL": "https://prd.example.com/odata/v2",
            "SF_TARGET_URL": "https://dev.example.com/odata/v2",
            "SF_SOURCE_USER": "prd_user",
            "SF_SOURCE_PASSWORD": "prd_pass",
            "SF_TARGET_USER": "dev_user",
            "SF_TARGET_PASSWORD": "dev_pass",
            "DRY_RUN": "true",
        }
        with patch.dict(os.environ, env, clear=False):
            cfg = config_from_env()
        self.assertEqual(cfg["prd"]["base_url"], "https://prd.example.com/odata/v2")
        self.assertEqual(cfg["prd"]["username"], "prd_user")
        self.assertEqual(cfg["dev"]["base_url"], "https://dev.example.com/odata/v2")
        self.assertTrue(cfg["options"]["dry_run"])


# ── sync_objects integration tests (all I/O mocked) ──────────────────────────

class TestSyncObjectsDryRun(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def _prd_record(self, code: str) -> Dict[str, Any]:
        return {
            "externalCode": code,
            "startDate": "/Date(946684800000)/",  # 2000-01-01
            "endDate": "/Date(253370764800000)/",  # 9999-12-31
            "status": "A",
            "country": "GB",
            "currency": "GBP",
            "name_defaultValue": f"Test {code}",
        }

    @patch("src.sync_engine.GapChecker")
    @patch("src.sync_engine.HierarchyResolver")
    @patch("src.sync_engine.SFClient")
    def test_dry_run_no_missing_returns_success(self, MockClient, MockResolver, MockGap):
        """When all entities exist in Dev, success=0 (nothing to create)."""
        path = _make_xlsx([("Department", "DEP-001")], self.tmp)

        # SFClient.get_entity_by_code returns a matching PRD record
        mock_prd_record = self._prd_record("DEP-001")
        mock_prd_record["startDate"] = "/Date(946684800000)/"

        prd_inst = MagicMock()
        dev_inst = MagicMock()
        MockClient.side_effect = [prd_inst, dev_inst]
        prd_inst.get_entity_by_code.return_value = [mock_prd_record]
        prd_inst.base_url = "https://prd.example.com/odata/v2"
        dev_inst.base_url = "https://dev.example.com/odata/v2"

        # HierarchyResolver.resolve returns a flat chain with just the dept
        resolver_inst = MagicMock()
        MockResolver.return_value = resolver_inst
        resolver_inst.resolve.return_value = [("Department", "DEP-001", mock_prd_record)]
        resolver_inst.get_cache.return_value = {}

        # GapChecker: entity already exists in Dev
        from src.gap_checker import GapCheckResult, DEV_EXISTS
        gap_inst = MagicMock()
        MockGap.return_value = gap_inst
        existing = GapCheckResult(status=DEV_EXISTS, entity_type="Department",
                                  entity_set="FODepartment", external_code="DEP-001")
        gap_inst.check_chain.return_value = [existing]
        gap_inst.get_results.return_value = {("Department", "DEP-001"): existing}

        result = sync_objects(
            source_config={"base_url": "https://prd.example.com/odata/v2",
                           "username": "u", "password": "p"},
            target_config={"base_url": "https://dev.example.com/odata/v2",
                           "username": "u", "password": "p"},
            input_file_path=path,
            dry_run=True,
            output_dir=self.tmp,
        )

        self.assertIsNone(result["error"])
        self.assertTrue(result["dry_run"])
        self.assertEqual(result["skipped"], 1)

    def test_no_valid_rows_returns_error(self):
        """Empty / all-invalid input returns error without crashing."""
        path = _make_xlsx([("Division", "DIV-001")], self.tmp)  # invalid type

        result = sync_objects(
            source_config={"base_url": "https://prd.example.com/odata/v2",
                           "username": "u", "password": "p"},
            target_config={"base_url": "https://dev.example.com/odata/v2",
                           "username": "u", "password": "p"},
            input_file_path=path,
            dry_run=True,
            output_dir=self.tmp,
        )

        self.assertIsNotNone(result["error"])

    def test_missing_file_returns_error(self):
        result = sync_objects(
            source_config={"base_url": "https://prd.example.com/odata/v2",
                           "username": "u", "password": "p"},
            target_config={"base_url": "https://dev.example.com/odata/v2",
                           "username": "u", "password": "p"},
            input_file_path="/nonexistent/input.xlsx",
            dry_run=True,
            output_dir=self.tmp,
        )
        self.assertIsNotNone(result["error"])
        self.assertIn("not found", result["error"].lower())


# ── Progress callback test ────────────────────────────────────────────────────

class TestProgressCallback(unittest.TestCase):

    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_callback_called_on_validation(self):
        """Progress callback receives at least the validation phase call."""
        path = _make_xlsx([("Division", "DIV-001")], self.tmp)  # will fail validation
        calls = []

        def _cb(phase, msg, pct):
            calls.append((phase, pct))

        sync_objects(
            source_config={"base_url": "https://prd.example.com/odata/v2",
                           "username": "u", "password": "p"},
            target_config={"base_url": "https://dev.example.com/odata/v2",
                           "username": "u", "password": "p"},
            input_file_path=path,
            dry_run=True,
            output_dir=self.tmp,
            progress_callback=_cb,
        )
        phases = [c[0] for c in calls]
        self.assertIn("init", phases)
        self.assertIn("validation", phases)
        # Percent should always be 0–100
        for _, pct in calls:
            self.assertGreaterEqual(pct, 0)
            self.assertLessEqual(pct, 100)


if __name__ == "__main__":
    unittest.main()
