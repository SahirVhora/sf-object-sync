import os
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.report_generator import _build_summary


def test_summary_counts_unique_entity_status():
    wb = openpyxl.Workbook()
    ws = wb.active

    run_meta = {
        "run_id": "run-123",
        "started_at": "2026-04-09T12:00:00+00:00",
        "completed_at": "2026-04-09T12:05:00+00:00",
        "config_file": "config.yaml",
        "input_file": "input.xlsx",
        "dry_run": False,
        "prd_url": "https://prd.example.com/odata/v2/",
        "dev_url": "https://dev.example.com/odata/v2/",
        "total_rows": 1,
        "valid_rows": 1,
    }

    audit_records = [
        {
            "status": "DEV_EXISTS",
            "object_type": "Sub Department",
            "external_code": "10000752",
            "phase": "gap_check",
        },
        {
            "status": "DEV_EXISTS",
            "object_type": "Sub Department",
            "external_code": "10000752",
            "phase": "upload",
        },
    ]

    _build_summary(ws, run_meta, audit_records)

    values = {ws.cell(row=i, column=1).value: ws.cell(row=i, column=2).value for i in range(2, 40)}
    assert values["Dev Exists"] == 1
