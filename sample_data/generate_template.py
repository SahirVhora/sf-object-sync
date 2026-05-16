#!/usr/bin/env python3
"""
generate_template.py - creates sample_data/foundation_objects_template.xlsx

Run from the sf_object_sync root:
  python sample_data/generate_template.py
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required:  pip install openpyxl")
    sys.exit(1)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(OUT_DIR, "foundation_objects_template.xlsx")

SAMPLE_ROWS = [
    ("Sub Department", "10000073"),
    ("Sub Department", "10000099"),
    ("Department", "10016236"),
    ("Department", "DEPT-001"),
]

HEADER_FILL = PatternFill("solid", fgColor="0A2540")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FILL = PatternFill("solid", fgColor="EAF2FF")
BORDER = Border(
    left=Side(style="thin", color="C0C8D8"),
    right=Side(style="thin", color="C0C8D8"),
    top=Side(style="thin", color="C0C8D8"),
    bottom=Side(style="thin", color="C0C8D8"),
)


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foundation Objects"

    # ── Instructions row ──────────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"] = (
        "sf_object_sync input template  |  "
        "Columns required: Object, Code  |  "
        "Object values: Sub Department, Department"
    )
    ws["A1"].font = Font(italic=True, color="5A6A80", size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Object", "Code", "Notes (optional - ignored by tool)"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = BORDER
    ws.row_dimensions[2].height = 24

    # ── Sample rows ───────────────────────────────────────────────────────────
    notes = [
        "Replace with your Sub Department externalCode",
        "Replace with your Sub Department externalCode",
        "Replace with your Department externalCode",
        "Hyphens and underscores allowed in codes",
    ]
    for i, (obj, code) in enumerate(SAMPLE_ROWS):
        row_num = i + 3
        for col, val in enumerate([obj, code, notes[i]], start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = EXAMPLE_FILL if row_num % 2 == 1 else PatternFill()
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_num].height = 20

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 46

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A3"

    wb.save(OUT_FILE)
    print(f"Template created: {OUT_FILE}")


if __name__ == "__main__":
    main()
